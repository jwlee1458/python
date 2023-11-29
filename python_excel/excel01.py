import openpyxl
from faker import Faker
import os

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet['A1'] = "이름"
worksheet['B1'] = "이메일"
worksheet['C1'] = "전화번호"

fake = Faker('ko_KR')

for row in range(2, 50):
    worksheet.cell(row=row, column=1, value=fake.name())
    worksheet.cell(row=row, column=2, value=fake.email())
    worksheet.cell(row=row, column=3, value=fake.phone_number())

workbook.save(os.path.join("python_excel", "member.xlsx"))