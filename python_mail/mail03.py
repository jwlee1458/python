import requests
from bs4 import BeautifulSoup
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication 
from dotenv import load_dotenv
import os
import openpyxl
from datetime import datetime
import schedule
import time
from dotenv import load_dotenv, find_dotenv

dotenv_path = find_dotenv(filename=".env", raise_error_if_not_found=True)
load_dotenv(dotenv_path)
SECRET_ID = os.getenv("SECRET_ID")
SECRET_PASS = os.getenv("SECRET_PASS")
MY_EMAIL = os.getenv("MY_EMAIL")
YOUR_EMAIL = os.getenv("YOUR_EMAIL")

def mail_sender(file_path, file_name, now):
        smtp = smtplib.SMTP('smtp.naver.com', 587)
        smtp.ehlo()
        smtp.starttls()

        smtp.login(SECRET_ID,SECRET_PASS)

        myemail = MY_EMAIL
        youremail = YOUR_EMAIL

        msg = MIMEMultipart()

        msg['Subject'] =f"보안 동향 {now} 정보입니다."
        msg['From'] = myemail
        msg['To'] = youremail

        text = """
        <html>
        <body>
        <h2>보안 동향 {} 정보입니다.</h2>
        </body>
        </html>
        """.format(now)

        contentPart = MIMEText(text, "html") 
        msg.attach(contentPart) 

        etc_file_path = file_path
        with open(etc_file_path, 'rb') as f : 
                etc_part = MIMEApplication( f.read() )
                etc_part.add_header('Content-Disposition','attachment', filename=file_name)
                msg.attach(etc_part)

        smtp.sendmail( myemail,youremail,msg.as_string() )
        smtp.quit()

def job():
        now = datetime.now().strftime("%Y-%m-%d")
        url = "https://www.malware-traffic-analysis.net/2023/index.html"

        headers = {
                'User-Agent': 'Mozilla/5.0',
                'Content-Type': 'text/html; charset=utf-8'
        }

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = "설명"
        worksheet['B1'] = "URL링크"

        req = requests.get(url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")

        tags = soup.select("#main_content > div.content > ul > li > a.main_menu")

        row = 2
        for tag in tags:
                tag_text = tag.text
                tag_href = f"https://www.malware-traffic-analysis.net/2023/{tag['href']}"
                worksheet.cell(row=row, column=1, value=tag_text)
                worksheet.cell(row=row, column=2, value=tag_href)
                row = row + 1

        workbook.save(os.path.join("python_mail", f'malware_{now}.xlsx'))
        #메일보내기
        mail_sender(os.path.join("python_mail", f'malware_{now}.xlsx'), f'malware_{now}.xlsx', now)

#스케줄링
schedule.every(5).seconds.do(job)

while True:
    schedule.run_pending()
    time.sleep(1)