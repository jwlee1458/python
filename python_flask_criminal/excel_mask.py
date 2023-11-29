import openpyxl
import os

# 엑셀 파일 열기
workbook = openpyxl.load_workbook(os.path.join("python_flask_criminal", "korean_data.xlsx"))
sheet = workbook.active

masked_workbook = openpyxl.Workbook()
masked_sheet = masked_workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    name, address, phone_number, email = row[0], row[1], row[2], row[3]
    username, domain = email.split('@')
    masked_email = '*' * len(username) + '@' + domain
    maksed_data= [name, address, phone_number,masked_email]
    masked_sheet.append(maksed_data)

masked_workbook.save(os.path.join("python_flask_criminal", "masked_excel.xlsx"))