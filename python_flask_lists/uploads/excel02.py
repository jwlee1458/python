import openpyxl
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import os

now = datetime.now().strftime("%Y-%m-%d")
print(now)

workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet['A1'] = "제목"
worksheet['B1'] = "링크"

#크롤링
url = "https://www.malware-traffic-analysis.net/2023/index.html"

headers = {
        'User-Agent': 'Mozilla/5.0',
        'Content-Type': 'text/html; charset=utf-8'
}

req = requests.get(url, headers=headers)
soup = BeautifulSoup(req.text, "lxml")
tags = soup.select("#main_content > div.content > ul > li > a.main_menu")

row = 2
for tag in tags:
    tag_text = tag.text
    tag_href = f"https://www.malware-traffic-analysis.net/2023/{tag['href']}"
    worksheet.cell(row=row, column=1, value=tag_text)
    worksheet.cell(row=row, column=2, value=tag_href)
    row = row + 1

workbook.save(os.path.join("python_excel", f"{now}_malwares.xlsx"))